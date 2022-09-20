import sys
import json
import numpy as np
import seaborn as sns
import openpyxl as pxl
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap

from pathExtract import TASK_ID, TRIAL_ID

try:
    TASK_ID = sys.argv[1]
except IndexError:
    pass
try:
    TRIAL_ID = sys.argv[2]
except IndexError:
    pass


def processData(data_dict: dict) -> list[list[int]]:
    """
    Function to process the input dictionaries to a 2D list for plotting
        - Args
            - `data_dict`: dictionary of behaviors (and confidence)
        - Returns
        - 2-D array for the matshow plot
    """
    
    # for the normal dictionary 
    if type(next(iter(data_dict.values()))) == str: 
        # this way the data type of the computer output can be anything other than "str"
        behaviorsList = []
        for angle in np.linspace(20, 70, 11, dtype=int):
            behaviorList_I = []
            for velocity in np.linspace(1.0, 7.0, 13):

                if data_dict[f"V{velocity}_A{angle}"] == "FS":
                    behaviorList_I.append(0)
                if data_dict[f"V{velocity}_A{angle}"] == "RO":
                    behaviorList_I.append(1)
                if data_dict[f"V{velocity}_A{angle}"] == "RC":
                    behaviorList_I.append(2)

            behaviorsList.append(behaviorList_I)

        return behaviorsList
    
    # extracting the confidence values 
    #todo the loaded data is now a key-value thing so the indexing should change
    else:
        behaviorsList = []
        confidenceList = []
        for angle in np.linspace(20, 70, 11, dtype=int):
            # initializing each row 
            behaviorList_I = []
            confidenceList_I = []
            for velocity in np.linspace(1.0, 7.0, 13):

                # adding the confidence
                confidenceList_I.append(data_dict[f"V{velocity}_A{angle}"]['confidence'])

                if data_dict[f"V{velocity}_A{angle}"]['behavior'] == "FS":
                    behaviorList_I.append(0)

                if data_dict[f"V{velocity}_A{angle}"]['behavior'] == "RO":
                    behaviorList_I.append(1)
                    
                if data_dict[f"V{velocity}_A{angle}"]['behavior'] == "RC":
                    behaviorList_I.append(2)

            behaviorsList.append(behaviorList_I)
            confidenceList.append(confidenceList_I)

        return behaviorsList, confidenceList 
        


def main():

    with open(f"behaviors/computer/Computer_Behavior_{TASK_ID}_{TRIAL_ID}.json", 'r') as readC:
        compDict = json.load(readC)

    # the human dictionary is probably not sorted
    try:
        with open(f"behaviors/human/Human_Behavior_{TASK_ID}.json", 'r') as file:
            humanDict = json.load(file)
    
    # if there is no json file to refer to, go to the old xlsx file
    except FileNotFoundError:
        wb_file = pxl.load_workbook(f"behaviors/human/{TASK_ID}_batchManager.xlsx") #only need relative path
        sheet = wb_file.active
        MAXROW = sheet.max_row
        humanDict = {}
        for i in range(2, MAXROW+1):
            iterations_H = sheet.cell(row = i, column = 1).value
            behavior_H = sheet.cell(row = i, column = 4).value

            humanDict[iterations_H] = behavior_H
    
    #finding the unmatching keys
    unmatching = []
    for key in compDict:
        if compDict[key]['behavior'] != humanDict[key]:
            unmatching.append(key)

    #extracting x and y axes from the humanDict dictionary
    y_axis = list(np.linspace(20, 70, 11, dtype=int)) # angle
    x_axis = list(np.linspace(1.0, 7.0, 13)) # vel
    
    total_iter = len(x_axis)*len(y_axis)

    #calling the function on the dictionaries
    humanBehavior = np.array(processData(humanDict))
    comp_decision = processData(compDict)
    
    compBehavior = np.array(comp_decision[0])
    confidence = np.array(comp_decision[1])

    # diffBehavior = np.logical_xor(humanBehavior, compBehavior)
    diffBehavior = humanBehavior != compBehavior

    fig, (
        (ax1,ax2),
        (ax3,ax4)
            ) = plt.subplots(2,2, figsize=(10,9.5), sharey=False)
    fig.suptitle(f"{TASK_ID} Behavior Comparison (matching rate: {(total_iter-len(unmatching))/total_iter:.3})", 
        fontweight = "bold",
        fontsize = 16,

    )


    #colormap parameters
    myColors = (
        (199/255,45/255,34/255,1.0), 
        (224/255,227/255,34/255,1.0), 
        (64/255,133/255,27/255,1.0)
    )
    colors = LinearSegmentedColormap.from_list('Custom', myColors, len(myColors))

    #colormap parameters for matching/unmatching
    myColorsBW = (
        (0.8,0.8,0.8,1), 
        (0,0,0,1)
    )
    colorsBW = LinearSegmentedColormap.from_list('Custom', myColorsBW, len(myColorsBW))

    #plotting
    ax1.set_title("Computer plot")
    ax1 = sns.heatmap(
        ax = ax1,
        data = compBehavior,
        yticklabels=y_axis,
        linewidths=2,
        cmap=colors,
        vmax=2, vmin=0
    )
    ax1.invert_yaxis()
    ax1.set_xticklabels(x_axis, rotation=45)
    ax1.set_xlabel("Velocities")
    ax1.set_ylabel("Angles")
    ax1.margins(1)

    colorbar1 = ax1.collections[0].colorbar
    colorbar1.set_ticks([0.33,1.0,1.66])
    colorbar1.set_ticklabels(["FS","RO","RC"])


    ax2.set_title("Human Plot")
    ax2 = sns.heatmap(
        ax = ax2,
        data = humanBehavior,
        yticklabels=y_axis,
        linewidths=2,
        cmap=colors,
        vmax=2, vmin=0
    )
    ax2.invert_yaxis()
    ax2.set_xticklabels(x_axis, rotation=45)
    ax2.set_xlabel("Velocities")
    ax2.set_ylabel("Angles")

    colorbar2 = ax2.collections[0].colorbar
    colorbar2.set_ticks([0.33,1.0,1.66])
    colorbar2.set_ticklabels(["FS","RO","RC"])

    ax3.set_title("Unmatching")
    ax3 = sns.heatmap(
        ax = ax3,
        data = diffBehavior,
        yticklabels=y_axis,
        linewidths=2,
        cmap=colorsBW,
        # vmax=2, vmin=0
    )
    ax3.invert_yaxis()
    ax3.set_xticklabels(x_axis, rotation=45)
    ax3.set_xlabel("Velocities")
    ax3.set_ylabel("Angles")

    colorbar3 = ax3.collections[0].colorbar
    colorbar3.set_ticks([0.25, 0.75])
    colorbar3.set_ticklabels(["Matching", "Unmatching"])
    
    ax4.set_title(f"Confidence (Avg = {np.average(confidence): .3f})")
    ax4 = sns.heatmap(
        ax = ax4,
        data = confidence,
        yticklabels=y_axis,
        linewidths=2,
        cmap="Reds",
        # vmax=2, vmin=0
    )
    ax4.invert_yaxis()
    ax4.set_xticklabels(x_axis, rotation=45)
    ax4.set_xlabel("Velocities")
    ax4.set_ylabel("Angles")

    colorbar4 = ax3.collections[0].colorbar
    colorbar4.set_ticks(np.linspace(0,1,11))
    

    outputText = f"Matching Rate: {(total_iter-len(unmatching))/total_iter:.3}\n{len(unmatching)} Unmatching parameters: {unmatching}"
    plt.subplots_adjust(
        top=0.88,
        bottom=0.21,
        left=0.13,
        right=0.9,
        hspace=0.31,
        wspace=0.2
    )
    plt.gcf().text(0.1, 0.1, outputText, fontsize=12)

    plt.savefig(f"output_plots/behavior_comparisons/behaviorPlot_{TASK_ID}_{TRIAL_ID}.svg", format = "svg")
    plt.show()
    # output_plots\behavior_comparisons
    
        

if __name__ == "__main__":
    main()
