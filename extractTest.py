import os
import sys
import json
import logging
import numpy as np
import openpyxl as pxl
from time import perf_counter

from post_processing import Comparer, Visualizer
from pathExtract import TASK_ID, TRIAL_ID
#todo   Implement task ID and trial ID to differentiate between tasks and such

try:
    TASK_ID = sys.argv[1]
except IndexError:
    pass
try:
    TRIAL_ID = sys.argv[2]
except IndexError:
    pass

def main():
    # loading the json file with human decisions for comparison
    # with open(f"behaviors/human/Human_Behavior_{TASK_ID}.json", 'r') as file:
    #     humanDict = json.load(file)

    try:
        with open(f"behaviors/human/Human_Behavior_{TASK_ID}.json", 'r') as file:
            humanDict = json.load(file)

    except FileNotFoundError:
        wb_file = pxl.load_workbook(f"behaviors/human/{TASK_ID}_batchManager.xlsx") #only need relative path
        sheet = wb_file.active
        MAXROW = sheet.max_row
        humanDict = {}
        for i in range(2, MAXROW+1):
            iterations_H = sheet.cell(row = i, column = 1).value
            behavior_H = sheet.cell(row = i, column = 4).value

            humanDict[iterations_H] = behavior_H

    # dictionary to hold the computer behavior classifications
    compDict = {}
    unmatching = []

    # vectors to loop angles -> velocities
    velocities = np.linspace(1.0, 7.0, 13)
    angles = np.linspace(20, 70, 11, dtype=int)
    # velocities = [7.0]
    # angles = [25]

    for angle in angles:
        for velocity in velocities:
            # the path for the data extracts for the computer decision
            path = f"data_Extracts/data_Extract_{TASK_ID}_{TRIAL_ID}/outputs_V{velocity}_A{angle}.json"

            #* creating an output object that holds the necessary data
            output = Comparer(path, humanDict[f"V{velocity}_A{angle}"])
            human_dec = output.behavior_obj.human_decision

            # information for conditionals and such all included for better analysis
            comp_dec = {
                "behavior" : output.behavior_obj.computer_decision,
                "confidence": output.behavior_obj.confidence,
                "contact_pIDs": output.behavior_obj.contact_pIDs,
                "airborne": output.behavior_obj.airborne,
                "surpasses": output.behavior_obj.surpasses,
                "crater_out": output.behavior_obj.crater_out
            }
            
            #* plotting the output behavior and saving the figure
            plotter = Visualizer(output.behavior_obj)

            save_path = f"output_plots/path_plots/path_plots_{TASK_ID}_{TRIAL_ID}"
            os.makedirs(save_path, exist_ok=True)
            
            # plotting funciton
            plotter.path_plots(save_path, "svg")

            #* saving the computer behavior
            #todo Export <<output.behavior_obj.decision>> this here
            #! comp_dec is a tuple --> (behavior, confidence)
            compDict[f"V{velocity}_A{angle}"] = comp_dec

            
            #* printing the behavior output
            if human_dec == comp_dec["behavior"]:
                behavior_str = f"Computer / Human: [{ comp_dec['behavior'] }/{human_dec}]\tConfidence: [{comp_dec['confidence']}]"
            else:
                # behavior_str = f"Computer / Human: [{comp_dec[0]}/{human_dec}]\tConfidence: [{comp_dec[1]}] <-- UNMATCHING" 
                behavior_str = f"Computer / Human: [{ comp_dec['behavior'] }/{human_dec}]\tConfidence: [{comp_dec['confidence']}] <-- UNMATCHING" 
                unmatching.append(f"V{velocity}_A{angle}") 

            logging.basicConfig( 
                level = logging.INFO, 
                format = "  %(message)s", 

                # outputting to both a log file and the stdout 
                handlers = [logging.FileHandler(f"output_logs/extract_{TASK_ID}_{TRIAL_ID}.log"), logging.StreamHandler(sys.stdout)]
            ) 
            logging.info(f"\n Iteration: {output.behavior_obj.iteration}\t{behavior_str}") 
            # printing the reasons line by line 
            for idx, reason in enumerate(output.behavior_obj.reasons): 
                logging.info(f"\tReason {idx}: {reason}") 

            # outputting the final vote 
            logging.info(f"\tVotes: {output.behavior_obj.decisionDict}")
                
    logging.info(f"\nMatch rate: { ( len(humanDict)-len(unmatching) )/len(humanDict):.3f}")    
    logging.info(f"{len(unmatching)} unmatching parameters: {unmatching}")    

    # printing the outputs as a table
    logging.info("\n%10s%10s%10s" %("iter", "human", "computer"))
    for key in unmatching:        
        logging.info("%10s%10s%10s" %(key, humanDict[key], compDict[key]["behavior"]))
    
    # saving the computer behavior classifications    
    with open(f"behaviors/computer/Computer_Behavior_{TASK_ID}_{TRIAL_ID}.json",'w') as file:        
        json.dump(compDict, file, indent=4)
        
if __name__ == "__main__":    
    start = perf_counter()    
    main()    
    end = perf_counter()    

    # showing the behavior plot
    import behaviorPlots
    behaviorPlots.main()

    print(f"Total runtime: { int( (end-start)//60 ) }:{(end-start)%60 : .2f}")
